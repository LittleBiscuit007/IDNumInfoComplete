import openpyxl as openpyxl

import conf


class LiangFengVillageInfo(object):
    """
    获取所有的组别和户主姓名
    获取各组的户主姓名和身份证，在各组全员信息表中
    写入辣椒表
    """
    def get_all_name(self):
        """
        获取所有的组别和户主姓名
        :return: [[category, name], ...]
        """
        all_name = []

        workbook = openpyxl.load_workbook(conf.ChiliBook)
        sheet_names = workbook.get_sheet_names()
        # 获取第一张表，可修改获取sheet_names的下标值获取不同的表
        ws = workbook.get_sheet_by_name(sheet_names[0])
        data = ws[conf.ChiliBookStartData:conf.ChiliBookEndData]

        for row in data:
            all_name.append([row[0].value, row[1].value])
            # print([row[0].value, row[1].value])
        return all_name

    def get_ID(self, bookname):
        """
        获取各组的户主姓名和身份证，在各组全员信息表中
        :return: {name: ID, ...}
        """
        ID_info = {}

        workbook = openpyxl.load_workbook(bookname)
        sheet_names = workbook.get_sheet_names()
        # 获取第一张表，可修改获取sheet_names的下标值获取不同的表
        ws = workbook.get_sheet_by_name(sheet_names[0])
        data = ws[conf.TaoJiaInfoBookStartData:conf.TaoJiaInfoBookEndData]

        for row in data:
            if row[0].value is None:
                continue
            ID_info[row[0].value] = row[3].value
            # print([row[0].value, row[3].value])
            # print(ID_info)
        return ID_info

    def get_phone(self, bookname):
        """
        获取所有人的姓名. 电话
        :return: {name: phone, ...}
        """
        phone_info = {}

        workbook = openpyxl.load_workbook(bookname)
        sheet_names = workbook.get_sheet_names()
        # 获取第一张表，可修改获取sheet_names的下标值获取不同的表
        ws = workbook.get_sheet_by_name(sheet_names[0])
        data = ws[conf.PhoneBookStartData:conf.PhoneBookEndData]

        for row in data:
            phone_info[row[0].value] = row[1].value
        return phone_info

    def write_info(self, all_name, all_phone, taojia_ID_info):
        """
        写入辣椒表
        :return:
        """
        row = conf.DataStartRow

        workbook = openpyxl.reader.excel.load_workbook(conf.ChiliBook)
        sheet_names = workbook.get_sheet_names()
        # 获取第一张表，可修改获取sheet_names的下标值获取不同的表
        ws = workbook.get_sheet_by_name(sheet_names[0])

        for a_name in all_name:
            # 写入电话信息
            try:
                ws[conf.PhoneCol + str(row)] = all_phone[a_name[1]]
            except KeyError:
                ws[conf.PhoneCol + str(row)] = "-"
            # 写入ID信息
            if a_name[0] == "桃家":
                try:
                    ws[conf.IDCol + str(row)] = taojia_ID_info[a_name[1]]
                except KeyError:
                    ws[conf.IDCol + str(row)] = "-"

            row += 1

        workbook.save(conf.ChiliBook)


if __name__ == '__main__':
    info = LiangFengVillageInfo()
    all_name = info.get_all_name()
    # print(all_name)
    taojia_ID_info = info.get_ID(conf.TaoJiaInfoBook)
    # print(taojia_ID_info)
    all_phone = info.get_phone(conf.PhoneBook)
    # print(all_phone)
    info.write_info(all_name, all_phone, taojia_ID_info)


